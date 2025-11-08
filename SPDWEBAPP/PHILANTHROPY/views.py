from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Philanthropy_Hours_Event_and_Request
from AUTHENTICATE.models import Brother_Profile, Role
from django.utils import timezone

from PARLEYPRO.pp_decorators import requires_role
import calendar




@login_required
@requires_role('ACTIVE')
def philanthropy_dashboard(request):
    #import functions to collect context
    from .views_functions import get_academic_year_dates, get_available_academic_years
    from .views_functions import get_user_events_and_total_hours, collect_philanthropy_statistics

    # Get selected academic year start (e.g. “2024” for 2024-2025 AY)
    selected_year = int(request.GET.get('year', 
        (timezone.now().year - 1) if timezone.now().month <= 6 else timezone.now().year
    ))
    start_date, end_date = get_academic_year_dates(selected_year)

    # Get all brothers with the 'Active Member' role
    active_brothers = Brother_Profile.objects.filter(
        roles__name='ACTIVE'
    ).select_related('user')
    
    brothers_progress, total_chapter_hours, members_meeting_goal = collect_philanthropy_statistics(active_brothers, start_date, end_date)
    # Calculate averages
    total_active_members = len(active_brothers)
    avg_hours_per_brother = total_chapter_hours / total_active_members if total_active_members > 0 else 0

    # Get user's personal events
    is_philanthropy_chair = request.user.brother_profile.roles.filter(name='PHILANTHROPY_CHAIR').exists()
    user_events = None
    user_total_hours = 0
    
    user_events, user_total_hours = get_user_events_and_total_hours(request,start_date, end_date)  
    context = {
        'brothers_progress': sorted(brothers_progress, key=lambda x: x['total_hours'], reverse=True),
        'total_chapter_hours': total_chapter_hours,
        'avg_hours_per_brother': avg_hours_per_brother,
        'members_meeting_goal': members_meeting_goal,
        'total_active_members': total_active_members,
        'user_events': user_events,
        'user_total_hours': user_total_hours,
        'is_philanthropy_chair': is_philanthropy_chair,
        'available_years': get_available_academic_years(),
        'selected_year': selected_year,
        'year_label': f"{selected_year}–{selected_year+1} Academic Year"
    }
    
    return render(request, 'philanthropy_dashboard.html', context)

@login_required
@requires_role('ACTIVE')
def philanthropy_request(request):
    if request.method == 'POST':
        try:
            hours = float(request.POST['hours'])
            if hours < 0.5:
                raise ValueError("Must request at least 0.5 hours.")
            
            new_event = Philanthropy_Hours_Event_and_Request(
                user=request.user,
                philanthropy_event_title=request.POST['event_title'],
                philanthropy_event_hours=hours,
                philanthropy_event_date=request.POST['event_date'],
                philanthropy_approval_status='requested'
            )
            new_event.save()
            messages.success(request, 'Philanthropy hours submitted successfully!')
            return redirect('philanthropy_dashboard')
        except Exception as e:
            messages.error(request, f'Error submitting hours: {str(e)}')
    
    return render(request, 'philanthropy_request.html')

@requires_role('PHIL_CHAIR','ACTIVE')
@login_required
def philanthropy_approve(request):
    from .views_functions import philanthropy_approval_POST
    
    if request.method == 'POST':
        philanthropy_approval_POST(request)
    # Get pending requests
    pending_requests = Philanthropy_Hours_Event_and_Request.objects.filter(
        philanthropy_approval_status='requested'
    ).select_related('user', 'user__brother_profile').order_by('-philanthropy_event_date')
    
    return render(request, 'philanthropy_approve.html', {'pending_requests': pending_requests})


@login_required
@requires_role('ACTIVE')
def brother_philanthropy_history(request, user_id):
    from .views_functions import retrieve_individual_brother_history
    try:
        brother = Brother_Profile.objects.select_related('user').get(user_id=user_id)
        is_philanthropy_chair = request.user.brother_profile.roles.filter(name="PHIL_CHAIR").exists()

        #grabs and sorts all of the brothers philanthropy history, outputs the total approved hours per semester
        academic_year_totals, lifetime_event_total = retrieve_individual_brother_history(brother, user_id)
        context = {
            'brother': brother,
            'academic_year_totals': academic_year_totals,
            'lifetime_total': lifetime_event_total,
            'is_philanthropy_chair': is_philanthropy_chair
        }

        
        return render(request, 'philanthropy_brother_history.html', context)
        
    except Brother_Profile.DoesNotExist:
        messages.error(request, 'Brother not found.')
        return redirect('philanthropy_dashboard')







# EXPORT PHILANTHROPY FUNCTION

import io
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

@login_required
#@requires_role('PHIL_CHAIR')
def export_approved_philanthropy_hours(request):
    # 1. Get all active brothers
    brothers = Brother_Profile.objects.filter(roles__name='ACTIVE').select_related('user')
    name_map = {bp.user_id: f"{bp.firstName} {bp.lastName}" for bp in brothers}
    
    # 2. Get approved events
    events = (Philanthropy_Hours_Event_and_Request.objects
        .filter(philanthropy_approval_status='approved')
        .select_related('user')
        .order_by('philanthropy_event_date')
    )

    # 3. Organize by user full name
    by_user = defaultdict(list)
    for e in events:
        name = name_map.get(e.user.id, e.user.username)
        by_user[name].append({
            'Date': e.philanthropy_event_date.strftime('%Y-%m-%d'),
            'Event': e.philanthropy_event_title,
            'Hours': e.philanthropy_event_hours,
            'Approved By': e.philanthropy_event_approver_name or '—',
        })

    # 4. Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = 'Philanthropy Hours'
    subheaders = ['Date', 'Event', 'Hours', 'Approved By']
    palette = ['FFC7CE','C6EFCE','FFEB9C','9BC2E6','D9D9D9']
    block_size = len(subheaders) + 1
    max_rows = max((len(events) for events in by_user.values()), default=0)

    # 5. Write each block
    for idx, member_name in enumerate(sorted(by_user)):
        col0 = idx * block_size + 1
        cols = [col0 + i for i in range(len(subheaders))]
        spacer = col0 + len(subheaders)
        total_hours = sum(item['Hours'] for item in by_user[member_name])
        if total_hours < 5:
            fill_color = 'FFC7CE'  # red
        elif total_hours < 10:
            fill_color = 'FFEB9C'  # yellow
        else:
            fill_color = 'C6EFCE'  # green
        fill = PatternFill('solid', fgColor=fill_color)



        # Header
        ws.merge_cells(start_row=1, start_column=col0, end_row=1, end_column=cols[-1])
        hcell = ws.cell(row=1, column=col0, value=member_name)
        hcell.fill = fill
        hcell.alignment = Alignment(horizontal='center', vertical='center')

        # Total hours
        ws.merge_cells(start_row=2, start_column=col0, end_row=2, end_column=cols[-1])
        tcell = ws.cell(row=2, column=col0, value=f"Total: {total_hours}")
        tcell.fill = fill
        tcell.alignment = Alignment(horizontal='center', vertical='center')

        # Subheaders
        for j, hdr in enumerate(subheaders):
            cell = ws.cell(row=3, column=cols[j], value=hdr)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for r, item in enumerate(by_user[member_name]):
            for j, key in enumerate(subheaders):
                ws.cell(row=4 + r, column=cols[j], value=item[key])

    # 6. Autosize
    for idx, _ in enumerate(by_user):
        base = idx * block_size + 1
        for j, hdr in enumerate(subheaders):
            col_idx = base + j
            letter = get_column_letter(col_idx)
            if hdr == 'Event':
                for row in range(3, 4 + max_rows):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True)
                ws.column_dimensions[letter].width = 15
            else:
                maxlen = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ''))
                    for r in range(1, 4 + max_rows)
                )
                ws.column_dimensions[letter].width = maxlen + 2
        # Spacer column
        spacer_col = base + len(subheaders)
        ws.column_dimensions[get_column_letter(spacer_col)].width = 2

    # 7. Response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=approved_philanthropy_hours.xlsx'
    return response

from django.shortcuts import get_object_or_404

@login_required
@requires_role('PHIL_CHAIR')
def philanthropy_edit_event(request, event_id):
    # Only PHIL_CHAIR can edit
    #if not check_user_role(request.user, 'PHIL_CHAIR'):
    #    messages.error(request, 'Only Philanthropy Chair members can edit events.')
    #    return redirect('philanthropy_dashboard')

    event = get_object_or_404(Philanthropy_Hours_Event_and_Request, id=event_id)

    if request.method == 'POST':
        event.philanthropy_event_title = request.POST.get('event_title')
        event.philanthropy_event_hours = float(request.POST.get('hours'))
        event.philanthropy_event_date = request.POST.get('event_date')

        event.last_edited_by = request.user
        event.last_edited_at = timezone.now()
        event.save()

        messages.success(request, 'Event updated successfully.')
        return redirect('brother_philanthropy_history', user_id=event.user.id)

    context = {
        'editing': True,
        'event': event
    }
    return render(request, 'philanthropy_request.html', context)
