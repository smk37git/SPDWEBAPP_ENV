from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import get_object_or_404
from .views_functions import *
from django.contrib.auth.models import User

def newmember_dashboard(request):
    return render(request, 'newmember_dashboard.html')

@login_required
def newmember_marks_dashboard(request):
    new_members = get_new_members()
    new_members_progress = calculate_member_marks(new_members)
    
    is_active = check_user_role(request.user, 'ACTIVE')
    is_new_member = check_user_role(request.user, 'NEW_MEMBER')
    user_submitted_marks = get_user_submitted_marks(request.user) if is_active else None

    # Calculate analytics based on user role
    if is_new_member:
        personal_stats = get_personal_stats(request.user)
        analytics = None
    else:
        analytics = get_active_member_analytics(new_members_progress)
        personal_stats = None
    
    context = {
        'new_members_progress': new_members_progress,
        'user_submitted_marks': user_submitted_marks,
        'is_active': is_active,
        'is_new_member_board': check_user_role(request.user, 'NM_BOARD'),
        'analytics': analytics,
        'personal_stats': personal_stats
    }
    
    template_name = 'ACTIVE_newmember_marks_dashboard.html' if is_active else 'newmember_marks_dashboard.html'
    return render(request, template_name, context)

@login_required
def newmember_request(request):
    if not check_user_role(request.user, 'ACTIVE'):
        messages.error(request, 'Only active members can submit marks.')
        return redirect('newmember_marks_dashboard')
        
    if request.method == 'POST':
        success, message = create_mark_requests(
            requesting_user=request.user,
            target_users=request.POST.getlist('target_users'),
            mark_value=int(request.POST['mark_value']),
            mark_reason=request.POST['mark_reason'],
            mark_date=request.POST['mark_date']
        )
        
        if success:
            messages.success(request, message)
        else:
            messages.error(request, message)
        return redirect('newmember_marks_dashboard')
    
    new_members = get_new_members()
    return render(request, 'newmember_marks_request.html', {'new_members': new_members})

@login_required
def newmember_approve(request):
    if not check_user_role(request.user, 'NM_BOARD'):
        messages.error(request, 'Only New Member Board members can approve marks.')
        return redirect('newmember_marks_dashboard')
        
    if request.method == 'POST':
        success, message = process_mark_approval(
            request, 
            request.POST.get('mark_id'), 
            request.POST.get('action')
        )
        if success:
            messages.success(request, message)
        else:
            messages.error(request, message)
    
    pending_requests = get_pending_mark_requests()
    
    return render(request, 'newmember_marks_approve.html', {
        'pending_requests': pending_requests
    })

@login_required
def newmember_mark_history(request, user_id):
    try:
        new_member = Brother_Profile.objects.select_related('user').get(
            user_id=user_id,
            roles__name='NEW_MEMBER'
        )
        
        marks, total_marks = get_member_mark_history(user_id)

        is_nm_board = request.user.brother_profile.roles.filter(name="NM_BOARD").exists()
        
        context = {
            'new_member': new_member,
            'marks': marks,
            'total_marks': total_marks,
            'is_negative': total_marks < 0,
            'is_nm_board': is_nm_board,
        }
        
        return render(request, 'newmember_marks_history.html', context)
        
    except Brother_Profile.DoesNotExist:
        messages.error(request, 'New Member not found.')
        return redirect('newmember_dashboard')



@login_required
def newmember_edit_mark(request, mark_id):
    # Only NM_BOARD can edit
    if not check_user_role(request.user, 'NM_BOARD'):
        messages.error(request, 'Only New Member Board members can edit marks.')
        return redirect('newmember_marks_dashboard')

    mark = get_object_or_404(NewMember_Mark_Event_and_Request, id=mark_id)

    if request.method == 'POST':
        mark.mark_event_title = request.POST.get('mark_reason')
        mark.mark_value = int(request.POST.get('mark_value'))
        mark.mark_event_date = request.POST.get('mark_date')

        new_target_id = request.POST.get('target_users')
        if new_target_id and str(mark.target_user.id) != new_target_id:
            try:
                new_target_user = User.objects.get(id=new_target_id)
                mark.target_user = new_target_user
            except User.DoesNotExist:
                messages.error(request, "Invalid target user selected.")
                return redirect('newmember_edit_mark', mark_id=mark.id)

        mark.last_edited_by = request.user
        mark.last_edited_at = timezone.now()
        mark.save()

        messages.success(request, 'Mark updated successfully.')
        return redirect('newmember_mark_history', user_id=mark.target_user.id)

    # Prepopulate form with mark data
    new_members = get_new_members()
    context = {
        'new_members': new_members,
        'editing': True,
        'mark': mark
    }
    return render(request, 'newmember_marks_request.html', context)

# Code for exporting the new member marks into a spreadsheet
import io
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import NewMember_Mark_Event_and_Request
from AUTHENTICATE.models import Brother_Profile

@login_required
def export_approved_newmember_marks(request):
    # 1. Find current new‑member IDs
    new_member_ids = Brother_Profile.objects.filter(
        roles__name='NEW_MEMBER'
    ).values_list('user_id', flat=True)

    # 2. Load full names for all profiles
    profiles = Brother_Profile.objects.select_related('user')
    name_map = {bp.user_id: f"{bp.firstName} {bp.lastName}" for bp in profiles}

    # 3. Gather approved marks by target full name
    marks = (NewMember_Mark_Event_and_Request.objects
        .filter(
            mark_approval_status='approved',
            target_user__id__in=new_member_ids
        )
        .select_related('target_user', 'requesting_user')
        .order_by('mark_event_date')      # ← oldest first
    )


    by_user = defaultdict(list)
    for m in marks:
        requester_name = name_map.get(m.requesting_user.id, m.requesting_user.username)
        target_name    = name_map.get(m.target_user.id,    m.target_user.username)
        by_user[target_name].append({
            'Date': m.mark_event_date.strftime('%Y-%m-%d'),
            'Requesting Member': requester_name,
            'Reason'      : m.mark_event_title,
            'Mark Value'       : m.mark_value,
        })

    max_rows = max((len(events) for events in by_user.values()), default=0)

    # 4. Create workbook & sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'New Member Marks'

    # 5. Color palette for each member block
    palette = ['FFC7CE','C6EFCE','FFEB9C','9BC2E6','D9D9D9']

    # 6. Layout rows
    HEADER_ROW = 1
    TOTAL_ROW  = 2
    SUBHDR_ROW = 3
    DATA_START = 4

    # 7. Define sub‑columns in desired order
    subheaders = ['Date','Requesting Member','Reason','Mark Value']




    block_size = len(subheaders) + 1    # 4 data cols + 1 spacer

    # 8. Write each member’s block
    for idx, member_name in enumerate(sorted(by_user)):
        col0 = idx * block_size + 1
        cols = [col0 + i for i in range(4)]
        spacer = col0 + 3
        fill    = PatternFill('solid', fgColor=palette[idx % len(palette)])

        # Member name header
        ws.merge_cells(start_row=HEADER_ROW, start_column=col0,
                       end_row=HEADER_ROW,   end_column=cols[-1])
        hcell = ws.cell(row=HEADER_ROW, column=col0, value=member_name)
        hcell.fill = fill
        hcell.alignment = Alignment(horizontal='center', vertical='center')

        # Total row
        total_points = sum(item['Mark Value'] for item in by_user[member_name])
        ws.merge_cells(start_row=TOTAL_ROW, start_column=col0,
                       end_row=TOTAL_ROW,   end_column=cols[-1])
        tcell = ws.cell(row=TOTAL_ROW, column=col0, value=f"Total: {total_points}")
        tcell.fill = fill
        tcell.alignment = Alignment(horizontal='center', vertical='center')

        # Sub‑headers
        for j, hdr in enumerate(subheaders):
            cell = ws.cell(row=SUBHDR_ROW, column=cols[j], value=hdr)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for r, event in enumerate(by_user[member_name]):
            for j, key in enumerate(subheaders):
                ws.cell(row=DATA_START + r, column=cols[j], value=event[key])

        # leave the 5th column blank

    # 9. Auto‑size & wrap
    for idx, member_name in enumerate(by_user):
        base = idx * block_size + 1

        # size each data column
        for j, hdr in enumerate(subheaders):
            col_idx = base + j
            letter  = get_column_letter(col_idx)

            # Set fixed widths
            if hdr == 'Reason':
                ws.column_dimensions[letter].width = 30
                for row in range(SUBHDR_ROW, DATA_START + max_rows):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True)
            elif hdr == 'Requesting Member':
                ws.column_dimensions[letter].width = 20
            elif hdr == 'Date':
                ws.column_dimensions[letter].width = 12
            elif hdr == 'Mark Value':
                ws.column_dimensions[letter].width = 10

        # spacer column
        spacer_col = base + len(subheaders)
        ws.column_dimensions[get_column_letter(spacer_col)].width = 2

    # 10. Output
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=approved_newmember_marks.xlsx'
    return response